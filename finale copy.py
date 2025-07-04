import pygame
import numpy as np
import pandas as pd
import random
import math
import time
from datetime import datetime, timedelta
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
from sklearn.preprocessing import LabelEncoder, StandardScaler
import threading
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from collections import deque, defaultdict

# Initialize Pygame
pygame.init()

# Constants
WINDOW_WIDTH = 1600
WINDOW_HEIGHT = 1000
GRID_SIZE = 6
CELL_SIZE = 100
GRID_OFFSET_X = 50
GRID_OFFSET_Y = 50
ROAD_WIDTH = 8

# Enhanced Colors with gradients
BLACK = (0, 0, 0)
WHITE = (255, 255, 255)
RED = (255, 0, 0)
GREEN = (0, 255, 0)
YELLOW = (255, 255, 0)
BLUE = (0, 0, 255)
GRAY = (128, 128, 128)
LIGHT_GRAY = (200, 200, 200)
DARK_GRAY = (64, 64, 64)
ORANGE = (255, 165, 0)
PURPLE = (128, 0, 128)
CYAN = (0, 255, 255)
BROWN = (139, 69, 19)
DEEP_BLUE = (25, 25, 112)
NEON_GREEN = (57, 255, 20)
NEON_BLUE = (31, 81, 255)
DARK_PURPLE = (48, 25, 52)
GOLD = (255, 215, 0)

# Gradient colors for modern UI
GRADIENT_DARK = (15, 15, 25)
GRADIENT_LIGHT = (35, 35, 55)
ACCENT_BLUE = (64, 156, 255)
ACCENT_GREEN = (0, 230, 118)
ACCENT_RED = (255, 69, 58)
ACCENT_ORANGE = (255, 159, 10)
ACCENT_PURPLE = (179,158,151)

class RLAgent:
    """Deep Q-Network agent for traffic light optimization"""
    def __init__(self, state_size=8, action_size=4, learning_rate=0.001):
        self.state_size = state_size
        self.action_size = action_size
        self.memory = deque(maxlen=10000)
        self.epsilon = 1.0  # exploration rate
        self.epsilon_min = 0.01
        self.epsilon_decay = 0.995
        self.learning_rate = learning_rate
        self.q_table = defaultdict(lambda: np.zeros(action_size))
        self.performance_history = deque(maxlen=1000)
        self.reward_history = deque(maxlen=100)
        self.learning_enabled = True
        
    def get_state(self, intersection, nearby_cars):
        """Convert intersection state to feature vector"""
        waiting_cars = len([car for car in nearby_cars if car.waiting])
        emergency_cars = len([car for car in nearby_cars if car.car_type == 'emergency'])
        avg_wait_time = np.mean([car.wait_time for car in nearby_cars if car.waiting]) if waiting_cars > 0 else 0
        congestion_level = min(10, waiting_cars)
        
        # Light state encoding
        light_state_encoding = {'red': 0, 'yellow': 1, 'green': 2}
        light_state = light_state_encoding[intersection.light_state]
        
        # Time features
        time_in_state = intersection.light_timer / 60.0  # Normalize
        
        # Traffic flow features
        cars_approaching = len([car for car in nearby_cars if not car.waiting and 
                              math.sqrt((car.x - intersection.x)**2 + (car.y - intersection.y)**2) < 80])
        
        # Performance metrics
        efficiency_score = intersection.total_cars_passed / max(1, intersection.light_changes)
        
        state = np.array([
            waiting_cars / 10.0,  # Normalize to 0-1
            emergency_cars,
            avg_wait_time / 100.0,  # Normalize
            congestion_level / 10.0,
            light_state / 2.0,
            time_in_state,
            cars_approaching / 10.0,
            efficiency_score / 10.0
        ])
        
        return tuple(state.round(2))  # Round for discrete state space
    
    def get_action(self, state):
        """Choose action using epsilon-greedy policy"""
        if not self.learning_enabled or np.random.rand() <= self.epsilon:
            return random.randrange(self.action_size)
        
        q_values = self.q_table[state]
        return np.argmax(q_values)
    
    def remember(self, state, action, reward, next_state, done):
        """Store experience in replay memory"""
        self.memory.append((state, action, reward, next_state, done))
    
    def calculate_reward(self, intersection, action, prev_state):
        """Calculate reward for the action taken"""
        reward = 0
        
        # Penalize excessive waiting
        total_wait_time = sum(car.wait_time for car in intersection.cars_waiting)
        reward -= total_wait_time * 0.1
        
        # Reward for handling emergency vehicles
        if intersection.emergency_present:
            if action == 2:  # Green light for emergency
                reward += 50
            else:
                reward -= 20
        
        # Reward for efficient traffic flow
        if intersection.vehicle_count > 0 and intersection.light_state == 'green':
            reward += min(10, intersection.vehicle_count * 2)
        
        # Penalize unnecessary light changes
        if action == 1:  # Force change
            reward -= 5
        
        # Reward for maintaining good flow
        if intersection.total_cars_passed > 0:
            efficiency = intersection.total_cars_passed / max(1, intersection.light_changes)
            reward += efficiency * 2
        
        # Penalize high congestion
        if intersection.vehicle_count > 8:
            reward -= 15
        
        return reward
    
    def replay(self, batch_size=32):
        """Train the model on a batch of experiences"""
        if len(self.memory) < batch_size:
            return
        
        batch = random.sample(self.memory, batch_size)
        
        for state, action, reward, next_state, done in batch:
            target = reward
            if not done:
                target += 0.95 * np.max(self.q_table[next_state])  # Gamma = 0.95
            
            # Q-learning update
            self.q_table[state][action] += self.learning_rate * (target - self.q_table[state][action])
        
        # Decay epsilon
        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay
        
        self.reward_history.append(np.mean([exp[2] for exp in batch]))

class Car:
    def __init__(self, x, y, direction, speed, target_intersection, car_type='normal'):
        self.start_x = x
        self.start_y = y
        self.x = x
        self.y = y
        self.direction = direction
        self.base_speed = speed
        self.speed = speed
        self.target_intersection = target_intersection
        self.car_type = car_type
        self.color = self.get_car_color()
        self.size = 16 if car_type == 'emergency' else 12
        self.waiting = False
        self.wait_time = 0
        self.total_wait_time = 0
        self.spawn_time = time.time()
        self.completed = False
        self.path = []
        self.current_path_index = 0
        self.glow_effect = 0
        
    def get_car_color(self):
        if self.car_type == 'emergency':
            return ACCENT_RED
        colors = [NEON_BLUE, PURPLE, CYAN, ACCENT_ORANGE, ACCENT_GREEN, GOLD]
        return random.choice(colors)
        
    def update(self, intersections):
        should_stop = False
        stop_distance = 35
        
        for intersection in intersections:
            distance = math.sqrt((self.x - intersection.x)**2 + (self.y - intersection.y)**2)
            
            if distance < stop_distance and intersection.light_state == 'red':
                if self.car_type != 'emergency':
                    approaching = False
                    if self.direction == 0 and self.y > intersection.y:
                        approaching = True
                    elif self.direction == 1 and self.x < intersection.x:
                        approaching = True
                    elif self.direction == 2 and self.y < intersection.y:
                        approaching = True
                    elif self.direction == 3 and self.x > intersection.x:
                        approaching = True
                    
                    if approaching:
                        should_stop = True
                        break
        
        if should_stop:
            self.waiting = True
            self.wait_time += 1
            self.total_wait_time += 1
            self.speed = 0
            self.glow_effect = (self.glow_effect + 0.2) % (2 * math.pi)
        else:
            self.waiting = False
            if self.wait_time > 0:
                self.wait_time = 0
            self.speed = self.base_speed
            self.glow_effect = 0
            
            # Move car
            if self.direction == 0:
                self.y -= self.speed
            elif self.direction == 1:
                self.x += self.speed
            elif self.direction == 2:
                self.y += self.speed
            elif self.direction == 3:
                self.x -= self.speed
            
    def draw(self, screen):
        # Glow effect for waiting cars
        if self.waiting and self.glow_effect > 0:
            glow_radius = int(self.size + 4 + 2 * math.sin(self.glow_effect))
            glow_color = (*self.color[:3], 50)  # Semi-transparent
            # Create glow surface
            glow_surf = pygame.Surface((glow_radius * 2, glow_radius * 2), pygame.SRCALPHA)
            pygame.draw.circle(glow_surf, (*self.color, 30), (glow_radius, glow_radius), glow_radius)
            screen.blit(glow_surf, (self.x - glow_radius, self.y - glow_radius))
        
        # Main car body with rounded corners
        car_rect = pygame.Rect(self.x - self.size//2, self.y - self.size//2, self.size, self.size)
        pygame.draw.rect(screen, self.color, car_rect, border_radius=5)
        
        # Car outline
        pygame.draw.rect(screen, WHITE, car_rect, 2, border_radius=5)
        
        # Emergency vehicle lights
        if self.car_type == 'emergency':
            if (time.time() * 6) % 1 < 0.5:
                pygame.draw.circle(screen, WHITE, (int(self.x - 4), int(self.y - 4)), 3)
                pygame.draw.circle(screen, ACCENT_RED, (int(self.x + 4), int(self.y + 4)), 3)
        
        # Direction indicator with modern arrow
        arrow_size = 8
        arrow_color = WHITE if not self.waiting else YELLOW
        
        if self.direction == 0:  # North
            points = [(self.x, self.y - arrow_size), 
                     (self.x - 5, self.y + 3), 
                     (self.x + 5, self.y + 3)]
        elif self.direction == 1:  # East
            points = [(self.x + arrow_size, self.y), 
                     (self.x - 3, self.y - 5), 
                     (self.x - 3, self.y + 5)]
        elif self.direction == 2:  # South
            points = [(self.x, self.y + arrow_size), 
                     (self.x - 5, self.y - 3), 
                     (self.x + 5, self.y - 3)]
        else:  # West
            points = [(self.x - arrow_size, self.y), 
                     (self.x + 3, self.y - 5), 
                     (self.x + 3, self.y + 5)]
        
        pygame.draw.polygon(screen, arrow_color, points)

class Intersection:
    def __init__(self, row, col, x, y):
        self.row = row
        self.col = col
        self.x = x
        self.y = y
        self.id = f"{chr(65 + row)}{col + 1}"
        self.light_state = 'red'
        self.light_timer = 0
        self.light_duration = 180
        self.cars_waiting = []
        self.congestion_level = 0
        self.vehicle_count = 0
        self.emergency_present = False
        self.total_cars_passed = 0
        self.avg_wait_time = 0
        self.light_changes = 0
        self.rl_agent = RLAgent()
        self.previous_state = None
        self.last_action = 0
        self.performance_score = 0
        self.animation_offset = 0
        
    def get_nearby_cars(self, cars):
        nearby = []
        for car in cars:
            distance = math.sqrt((car.x - self.x)**2 + (car.y - self.y)**2)
            if distance < 60:
                nearby.append(car)
        return nearby
        
    def update(self, cars, learning_enabled=True):
        self.light_timer += 1
        self.animation_offset = (self.animation_offset + 0.1) % (2 * math.pi)
        
        # Get nearby cars and update metrics
        nearby_cars = self.get_nearby_cars(cars)
        waiting_cars = [car for car in nearby_cars if car.waiting]
        
        self.cars_waiting = waiting_cars
        self.vehicle_count = len(waiting_cars)
        self.emergency_present = any(car.car_type == 'emergency' for car in nearby_cars)
        self.congestion_level = min(10, len(waiting_cars))
        
        if waiting_cars:
            self.avg_wait_time = sum(car.wait_time for car in waiting_cars) / len(waiting_cars)
        
        # RL-based traffic light control
        current_state = self.rl_agent.get_state(self, nearby_cars)
        
        if self.previous_state is not None and learning_enabled:
            # Calculate reward and learn
            reward = self.rl_agent.calculate_reward(self, self.last_action, self.previous_state)
            self.rl_agent.remember(self.previous_state, self.last_action, reward, current_state, False)
            
            # Train the agent
            if len(self.rl_agent.memory) > 100:
                self.rl_agent.replay(32)
        
        # Get action from RL agent
        action = self.rl_agent.get_action(current_state)
        self.last_action = action
        self.previous_state = current_state
        
        # Execute action
        self.execute_action(action)
        
        # Update performance score
        self.update_performance_score()
        
    def execute_action(self, action):
        """Execute the action chosen by RL agent"""
        # Actions: 0=maintain, 1=force_change, 2=extend_green, 3=quick_change
        
        if action == 1:  # Force change
            self.force_light_change()
        elif action == 2 and self.light_state == 'green':  # Extend green
            self.light_duration = min(self.light_duration + 60, 400)
        elif action == 3:  # Quick change
            self.light_timer = max(self.light_timer, self.light_duration - 30)
        
        # Normal light cycle with dynamic timing
        base_durations = self.get_dynamic_durations()
        
        if self.light_timer >= self.light_duration:
            if self.light_state == 'red':
                self.light_state = 'green'
                self.light_duration = base_durations['green']
                self.light_changes += 1
            elif self.light_state == 'green':
                self.light_state = 'yellow'
                self.light_duration = base_durations['yellow']
                self.light_changes += 1
            else:  # yellow
                self.light_state = 'red'
                self.light_duration = base_durations['red']
                self.light_changes += 1
            self.light_timer = 0
    
    
    def get_dynamic_durations(self):
        """Calculate dynamic light durations based on conditions"""
        base_green = 240
        base_red = 180
        base_yellow = 60
        
        # Emergency override
        if self.emergency_present:
            return {'green': 360, 'red': 60, 'yellow': 40}
        
        # Adjust based on congestion
        if self.vehicle_count > 8:
            base_green = 420
        elif self.vehicle_count > 5:
            base_green = 300
        elif self.vehicle_count == 0:
            base_green = 120
            base_red = 300
        
        return {'green': base_green, 'red': base_red, 'yellow': base_yellow}
    
    def force_light_change(self):
        """Force immediate light change"""
        if self.light_state == 'red':
            self.light_state = 'green'
            self.light_duration = self.get_dynamic_durations()['green']
        elif self.light_state == 'green':
            self.light_state = 'yellow'
            self.light_duration = self.get_dynamic_durations()['yellow']
        else:
            self.light_state = 'red' 
            self.light_duration = self.get_dynamic_durations()['red']
        
        self.light_timer = 0
        self.light_changes += 1
    
    def update_performance_score(self):
        """Calculate intersection performance score"""
        efficiency = self.total_cars_passed / max(1, self.light_changes)
        wait_penalty = self.avg_wait_time / 100.0
        congestion_penalty = self.vehicle_count / 10.0
        
        self.performance_score = max(0, efficiency - wait_penalty - congestion_penalty)
        
    def draw(self, screen, font):
        # Animated intersection base with gradient effect
        base_radius = 22
        glow_radius = int(base_radius + 3 * math.sin(self.animation_offset))
        
        # Glow effect
        glow_surf = pygame.Surface((glow_radius * 2, glow_radius * 2), pygame.SRCALPHA)
        pygame.draw.circle(glow_surf, (*DARK_GRAY, 100), (glow_radius, glow_radius), glow_radius)
        screen.blit(glow_surf, (self.x - glow_radius, self.y - glow_radius))
        
        # Main intersection
        pygame.draw.circle(screen, GRADIENT_LIGHT, (int(self.x), int(self.y)), base_radius)
        pygame.draw.circle(screen, LIGHT_GRAY, (int(self.x), int(self.y)), base_radius, 2)
        
        # Traffic light with modern design
        light_colors = {'red': ACCENT_RED, 'yellow': ACCENT_ORANGE, 'green': ACCENT_GREEN}
        light_color = light_colors[self.light_state]
        light_radius = 14
        
        # Light glow
        light_glow_surf = pygame.Surface((light_radius * 3, light_radius * 3), pygame.SRCALPHA)
        pygame.draw.circle(light_glow_surf, (*light_color, 80), (light_radius * 1.5, light_radius * 1.5), light_radius * 1.5)
        screen.blit(light_glow_surf, (self.x - light_radius * 1.5, self.y - light_radius * 1.5))
        
        # Main light
        pygame.draw.circle(screen, light_color, (int(self.x), int(self.y)), light_radius)
        pygame.draw.circle(screen, WHITE, (int(self.x), int(self.y)), light_radius, 3)
        
        # Intersection ID with modern styling
        id_surface = font.render(self.id, True, WHITE)
        id_rect = id_surface.get_rect(center=(self.x, self.y - 45))
        
        # ID background
        bg_rect = id_rect.inflate(12, 6)
        pygame.draw.rect(screen, GRADIENT_DARK, bg_rect, border_radius=8)
        pygame.draw.rect(screen, ACCENT_BLUE, bg_rect, 2, border_radius=8)
        screen.blit(id_surface, id_rect)
        
        # Performance indicator
        if self.vehicle_count > 0:
            perf_color = ACCENT_GREEN if self.performance_score > 5 else ACCENT_ORANGE if self.performance_score > 2 else ACCENT_RED
            perf_radius = int(12 + self.performance_score * 2)
            
            pygame.draw.circle(screen, perf_color, (int(self.x + 30), int(self.y - 30)), perf_radius, 4)
            
            # Vehicle count
            count_text = font.render(str(self.vehicle_count), True, WHITE)
            screen.blit(count_text, (self.x + 25, self.y - 35))

class TrafficSimulation:
    def __init__(self):
        self.screen = pygame.display.set_mode((WINDOW_WIDTH, WINDOW_HEIGHT))
        pygame.display.set_caption("AI Traffic Management System - Reinforcement Learning")
        self.clock = pygame.time.Clock()
        self.font = pygame.font.Font(None, 18)
        self.big_font = pygame.font.Font(None, 28)
        self.title_font = pygame.font.Font(None, 36)
        
        # Create intersections
        self.intersections = []
        for row in range(GRID_SIZE):
            for col in range(GRID_SIZE):
                x = GRID_OFFSET_X + col * CELL_SIZE + CELL_SIZE // 2
                y = GRID_OFFSET_Y + row * CELL_SIZE + CELL_SIZE // 2
                intersection = Intersection(row, col, x, y)
                self.intersections.append(intersection)
        
        self.cars = []
        self.data_logger = DataLogger()
        self.running = True
        self.paused = False
        self.simulation_speed = 60  # FPS
        self.speed_multiplier = 1.0
        self.frame_count = 0
        self.learning_enabled = True
        
        # Enhanced statistics
        self.total_cars_spawned = 0
        self.total_wait_time = 0
        self.emergency_vehicles = 0
        self.system_efficiency = 0
        self.avg_performance = 0
        self.learning_progress = 0
        # Performance tracking
        self.performance_history = deque(maxlen=100)
        self.efficiency_history = deque(maxlen=100)
        self.learning_history = deque(maxlen=100)
        self.reward_history = deque(maxlen=100)
        
        # UI state
        self.show_rl_info = True
        self.show_performance = True
        self.show_live_graphs = True
        
    def spawn_car(self):
        if len(self.cars) > 50:
            return
        
        # Enhanced spawn logic with better distribution
        spawn_points = []
        
        # Edge spawns with better positioning
        for i in range(GRID_SIZE):
            x = GRID_OFFSET_X + i * CELL_SIZE + CELL_SIZE // 2
            y = GRID_OFFSET_Y - 30
            spawn_points.append((x, y, 2))
            
            x = GRID_OFFSET_X + i * CELL_SIZE + CELL_SIZE // 2
            y = GRID_OFFSET_Y + GRID_SIZE * CELL_SIZE + 30
            spawn_points.append((x, y, 0))
            
            x = GRID_OFFSET_X - 30
            y = GRID_OFFSET_Y + i * CELL_SIZE + CELL_SIZE // 2
            spawn_points.append((x, y, 1))
            
            x = GRID_OFFSET_X + GRID_SIZE * CELL_SIZE + 30
            y = GRID_OFFSET_Y + i * CELL_SIZE + CELL_SIZE // 2
            spawn_points.append((x, y, 3))
        
        x, y, direction = random.choice(spawn_points)
        
        # Dynamic car type distribution
        car_type = 'emergency' if random.random() < 0.06 else 'normal'
        speed = random.uniform(1.8, 3.0) if car_type == 'normal' else random.uniform(3.0, 4.0)
        speed *= self.speed_multiplier
        
        target = random.choice(self.intersections)
        car = Car(x, y, direction, speed, target, car_type)
        self.cars.append(car)
        self.total_cars_spawned += 1
        
        if car_type == 'emergency':
            self.emergency_vehicles += 1
    
    def update_simulation(self):
        # Update intersections with RL
        for intersection in self.intersections:
            intersection.update(self.cars, self.learning_enabled)
        
        # Update cars
        cars_to_remove = []
        for car in self.cars:
            car.update(self.intersections)
            
            if (car.x < -100 or car.x > WINDOW_WIDTH + 100 or 
                car.y < -100 or car.y > WINDOW_HEIGHT + 100):
                car.completed = True
                self.data_logger.log_car(car)
                cars_to_remove.append(car)
        
        for car in cars_to_remove:
            self.cars.remove(car)
        
        # Update system metrics
        self.update_system_metrics()
    
    

    
    
    def draw_gradient_background(self):
        """Draw modern gradient background"""
        for y in range(WINDOW_HEIGHT):
            ratio = y / WINDOW_HEIGHT
            r = int(GRADIENT_DARK[0] + (GRADIENT_LIGHT[0] - GRADIENT_DARK[0]) * ratio)
            g = int(GRADIENT_DARK[1] + (GRADIENT_LIGHT[1] - GRADIENT_DARK[1]) * ratio)
            b = int(GRADIENT_DARK[2] + (GRADIENT_LIGHT[2] - GRADIENT_DARK[2]) * ratio)
            pygame.draw.line(self.screen, (r, g, b), (0, y), (WINDOW_WIDTH, y))
    
    def draw_modern_roads(self):
        """Draw roads with modern styling"""
        # Horizontal roads
        for i in range(GRID_SIZE):
            y = GRID_OFFSET_Y + i * CELL_SIZE + CELL_SIZE // 2
            # Road base
            pygame.draw.line(self.screen, DARK_GRAY, (0, y), (WINDOW_WIDTH, y), ROAD_WIDTH * 2 + 4)
            pygame.draw.line(self.screen, LIGHT_GRAY, (0, y), (WINDOW_WIDTH, y), ROAD_WIDTH * 2)
            # Animated center line
            dash_length = 20
            gap_length = 15
            offset = (self.frame_count // 2) % (dash_length + gap_length)
            
            for x in range(-offset, WINDOW_WIDTH + dash_length, dash_length + gap_length):
                if x >= 0 and x < WINDOW_WIDTH:
                    pygame.draw.line(self.screen, YELLOW, (x, y), (min(x + dash_length, WINDOW_WIDTH), y), 2)
        
        # Vertical roads
        for i in range(GRID_SIZE):
            x = GRID_OFFSET_X + i * CELL_SIZE + CELL_SIZE // 2
            pygame.draw.line(self.screen, DARK_GRAY, (x, 0), (x, WINDOW_HEIGHT), ROAD_WIDTH * 2 + 4)
            pygame.draw.line(self.screen, LIGHT_GRAY, (x, 0), (x, WINDOW_HEIGHT), ROAD_WIDTH * 2)
            
            offset = (self.frame_count // 2) % 35
            for y in range(-offset, WINDOW_HEIGHT + 20, 35):
                if y >= 0 and y < WINDOW_HEIGHT:
                    pygame.draw.line(self.screen, YELLOW, (x, y), (x, min(y + 20, WINDOW_HEIGHT)), 2)
    
    # Complete the draw_modern_ui method and add missing components

    def draw_modern_ui(self):
        """Draw enhanced modern UI"""
        # Main control panel
        panel_width = 380  # Slightly smaller
        panel_height = WINDOW_HEIGHT - 100
        panel_x = WINDOW_WIDTH - panel_width - 30  # More margin
        panel_y = 20
        
        # Panel background with gradient
        panel_rect = pygame.Rect(panel_x, panel_y, panel_width, panel_height)
        
        # Create gradient surface
        panel_surf = pygame.Surface((panel_width, panel_height), pygame.SRCALPHA)
        for y in range(panel_height):
            alpha = 180 - int(30 * y / panel_height)
            color = (*GRADIENT_DARK, alpha)
            pygame.draw.line(panel_surf, color, (0, y), (panel_width, y))
        
        self.screen.blit(panel_surf, (panel_x, panel_y))
        pygame.draw.rect(self.screen, ACCENT_BLUE, panel_rect, 3, border_radius=15)
        
        # Title
        title_text = self.title_font.render("AI Traffic Control", True, WHITE)
        self.screen.blit(title_text, (panel_x + 20, panel_y + 20))
        
        # System Status
        y_offset = panel_y + 70
        self.draw_stat_item("System Status", f"{'Learning' if self.learning_enabled else 'Manual'}", 
                           panel_x + 20, y_offset, ACCENT_GREEN if self.learning_enabled else ACCENT_ORANGE)
        
        y_offset += 40
        self.draw_stat_item("Cars Active", str(len(self.cars)), panel_x + 20, y_offset, ACCENT_BLUE)
        
        y_offset += 30
        self.draw_stat_item("Total Spawned", str(self.total_cars_spawned), panel_x + 20, y_offset, WHITE)
        
        y_offset += 30
        self.draw_stat_item("Emergency Vehicles", str(self.emergency_vehicles), panel_x + 20, y_offset, ACCENT_RED)
        
        y_offset += 30
        self.draw_stat_item("System Efficiency", f"{self.system_efficiency:.1f}%", 
                           panel_x + 20, y_offset, ACCENT_GREEN if self.system_efficiency > 70 else ACCENT_ORANGE)
        
        y_offset += 30
        self.draw_stat_item("Avg Performance", f"{self.avg_performance:.2f}", panel_x + 20, y_offset, CYAN)
        
        y_offset += 30
        self.draw_stat_item("Learning Progress", f"{self.learning_progress:.1f}%", 
                           panel_x + 20, y_offset, NEON_GREEN)
        
        # Performance graph
        y_offset += 60
        if self.show_performance:
            self.draw_performance_graph(panel_x + 20, y_offset, panel_width - 40, 120)
        
        # RL Information
        y_offset += 150
        if self.show_rl_info:
            self.draw_rl_info(panel_x + 20, y_offset, panel_width - 40)
        
        # Controls
        y_offset += 200
        self.draw_controls(panel_x + 20, y_offset)
        
        # Status indicators
        self.draw_status_indicators()
    
    def draw_stat_item(self, label, value, x, y, color):
        """Draw a modern stat item"""
        label_surf = self.font.render(label, True, LIGHT_GRAY)
        value_surf = self.big_font.render(value, True, color)
        
        self.screen.blit(label_surf, (x, y))
        self.screen.blit(value_surf, (x + 200, y - 3))
    
def draw_live_graphs(self):
    """Draw live learning status graphs"""
    if not self.show_live_graphs:
        return
    
    # Position graphs in bottom left
    graph_width = 300
    graph_height = 180
    graph_x = 20
    graph_y = WINDOW_HEIGHT - graph_height - 40
    
    # Background for graphs area
    graphs_rect = pygame.Rect(graph_x - 10, graph_y - 30, graph_width + 20, graph_height + 60)
    pygame.draw.rect(self.screen, (*GRADIENT_DARK, 200), graphs_rect, border_radius=12)
    pygame.draw.rect(self.screen, ACCENT_PURPLE, graphs_rect, 2, border_radius=12)
    
    # Title
    title_surf = self.big_font.render("Live Learning Analytics", True, WHITE)
    self.screen.blit(title_surf, (graph_x, graph_y - 25))
    
    # Draw four mini graphs
    mini_height = (graph_height - 30) // 2
    mini_width = (graph_width - 10) // 2
    
    # Efficiency graph (top-left)
    self.draw_mini_graph(graph_x, graph_y, mini_width, mini_height, 
                        list(self.efficiency_history), "Efficiency %", ACCENT_GREEN)
    
    # Learning progress (top-right)
    self.draw_mini_graph(graph_x + mini_width + 10, graph_y, mini_width, mini_height,
                        list(self.learning_history), "Learning %", NEON_BLUE)
    
    # Performance (bottom-left)
    self.draw_mini_graph(graph_x, graph_y + mini_height + 10, mini_width, mini_height,
                        list(self.performance_history), "Performance", CYAN)
    
    # Rewards (bottom-right)
    self.draw_mini_graph(graph_x + mini_width + 10, graph_y + mini_height + 10, 
                        mini_width, mini_height, list(self.reward_history), "Avg Reward", ACCENT_ORANGE)

def draw_mini_graph(self, x, y, width, height, data, title, color):
    """Draw a mini graph for live analytics"""
    # Background
    pygame.draw.rect(self.screen, (*DARK_GRAY, 100), (x, y, width, height), border_radius=5)
    pygame.draw.rect(self.screen, color, (x, y, width, height), 1, border_radius=5)
    
    # Title
    title_surf = pygame.font.Font(None, 16).render(title, True, color)
    self.screen.blit(title_surf, (x + 5, y + 2))
    
    if len(data) > 1:
        # Scale data
        max_val = max(data) if data else 1
        min_val = min(data) if data else 0
        if max_val == min_val:
            max_val = min_val + 1
        
        graph_height = height - 25
        graph_width = width - 10
        
        # Draw line
        points = []
        for i, value in enumerate(data[-20:]):  # Last 20 points
            px = x + 5 + (i * graph_width) // max(len(data[-20:]) - 1, 1)
            py = y + 20 + graph_height - int((value - min_val) / (max_val - min_val) * graph_height)
            points.append((px, py))
        
        if len(points) > 1:
            pygame.draw.lines(self.screen, color, False, points, 2)
        
        # Current value
        if data:
            current_surf = pygame.font.Font(None, 14).render(f"{data[-1]:.1f}", True, WHITE)
            self.screen.blit(current_surf, (x + width - 35, y + height - 18))
    
    def draw_rl_info(self, x, y, width):
        """Draw RL agent information"""
        # RL Info background
        info_rect = pygame.Rect(x, y, width, 180)
        pygame.draw.rect(self.screen, (*GRADIENT_DARK, 150), info_rect, border_radius=8)
        pygame.draw.rect(self.screen, ACCENT_ORANGE, info_rect, 2, border_radius=8)
        
        # Title
        title_surf = self.font.render("Reinforcement Learning Status", True, WHITE)
        self.screen.blit(title_surf, (x + 10, y + 5))
        
        y_pos = y + 30
        
        # Average epsilon across all intersections
        avg_epsilon = np.mean([i.rl_agent.epsilon for i in self.intersections])
        epsilon_surf = self.font.render(f"Exploration Rate: {avg_epsilon:.3f}", True, CYAN)
        self.screen.blit(epsilon_surf, (x + 10, y_pos))
        
        y_pos += 25
        memory_size = np.mean([len(i.rl_agent.memory) for i in self.intersections])
        memory_surf = self.font.render(f"Avg Memory Size: {memory_size:.0f}", True, CYAN)
        self.screen.blit(memory_surf, (x + 10, y_pos))
        
        y_pos += 25
        total_q_states = sum(len(i.rl_agent.q_table) for i in self.intersections)
        states_surf = self.font.render(f"Total Q-States: {total_q_states}", True, CYAN)
        self.screen.blit(states_surf, (x + 10, y_pos))
        
        y_pos += 25
        if self.intersections and self.intersections[0].rl_agent.reward_history:
            avg_reward = np.mean(list(self.intersections[0].rl_agent.reward_history))
            reward_surf = self.font.render(f"Avg Reward: {avg_reward:.2f}", True, CYAN)
            self.screen.blit(reward_surf, (x + 10, y_pos))
        
        # Learning progress bar
        y_pos += 35
        progress_width = width - 20
        progress_height = 12
        progress_rect = pygame.Rect(x + 10, y_pos, progress_width, progress_height)
        pygame.draw.rect(self.screen, DARK_GRAY, progress_rect, border_radius=6)
        
        progress_fill = int(progress_width * (self.learning_progress / 100))
        if progress_fill > 0:
            fill_rect = pygame.Rect(x + 10, y_pos, progress_fill, progress_height)
            pygame.draw.rect(self.screen, NEON_GREEN, fill_rect, border_radius=6)
        
        progress_text = self.font.render(f"Learning: {self.learning_progress:.1f}%", True, WHITE)
        self.screen.blit(progress_text, (x + 10, y_pos + 20))
    
def draw_controls(self, x, y):
    """Draw comprehensive control information"""
    controls = [
        ("SPACE", "Pause/Resume"),
        ("L", "Toggle AI Learning"),
        ("R", "Reset Simulation"),
        ("P", "Toggle Performance Panel"),
        ("I", "Toggle RL Info Panel"),
        ("G", "Toggle Live Graphs"),
        ("S", "Save Data Manually"),
        ("C", "Clear History"),
        ("1", "Spawn Emergency Vehicle"),
        ("2", "Spawn 5 Cars"),
        ("+/-", "Speed Control"),
        ("ESC", "Exit")
    ]
    
    title_surf = self.big_font.render("Controls:", True, ACCENT_BLUE)
    self.screen.blit(title_surf, (x, y - 25))
    
    for i, (key, description) in enumerate(controls):
        key_color = ACCENT_ORANGE
        desc_color = LIGHT_GRAY
        
        key_surf = self.font.render(key, True, key_color)
        desc_surf = self.font.render(f"- {description}", True, desc_color)
        
        self.screen.blit(key_surf, (x, y + i * 20))
        self.screen.blit(desc_surf, (x + 60, y + i * 20))
    
    def draw_status_indicators(self):
        """Draw system status indicators"""
        # Top status bar
        status_rect = pygame.Rect(20, 20, WINDOW_WIDTH - 480, 60)
        pygame.draw.rect(self.screen, (*GRADIENT_DARK, 200), status_rect, border_radius=12)
        pygame.draw.rect(self.screen, ACCENT_BLUE, status_rect, 2, border_radius=12)
        
        # Status text
        status_text = f"FPS: {self.clock.get_fps():.0f} | Speed: {self.speed_multiplier:.1f}x"
        if self.paused:
            status_text += " | PAUSED"
        
        status_surf = self.big_font.render(status_text, True, WHITE)
        self.screen.blit(status_surf, (30, 35))
        
        # Learning indicator
        if self.learning_enabled:
            indicator_rect = pygame.Rect(WINDOW_WIDTH - 500, 30, 20, 20)
            pygame.draw.circle(self.screen, NEON_GREEN, indicator_rect.center, 10)
            learn_text = self.font.render("AI Learning Active", True, NEON_GREEN)
            self.screen.blit(learn_text, (WINDOW_WIDTH - 470, 32))
    
def handle_events(self):
    """Handle pygame events with all controls"""
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            self.running = False
        elif event.type == pygame.KEYDOWN:
            if event.key == pygame.K_SPACE:
                self.paused = not self.paused
            elif event.key == pygame.K_l:
                self.learning_enabled = not self.learning_enabled
            elif event.key == pygame.K_r:
                self.reset_simulation()
            elif event.key == pygame.K_p:
                self.show_performance = not self.show_performance
            elif event.key == pygame.K_i:
                self.show_rl_info = not self.show_rl_info
            elif event.key == pygame.K_g:  # NEW - Toggle live graphs
                self.show_live_graphs = not self.show_live_graphs
            elif event.key == pygame.K_s:  # NEW - Save data manually
                self.data_logger.save_data()
                print("Data saved manually!")
            elif event.key == pygame.K_c:  # NEW - Clear performance history
                self.performance_history.clear()
                self.efficiency_history.clear()
                self.learning_history.clear()
                self.reward_history.clear()
                print("Performance history cleared!")
            elif event.key == pygame.K_PLUS or event.key == pygame.K_EQUALS:
                self.speed_multiplier = min(5.0, self.speed_multiplier + 0.2)
            elif event.key == pygame.K_MINUS:
                self.speed_multiplier = max(0.1, self.speed_multiplier - 0.2)
            elif event.key == pygame.K_1:  # NEW - Spawn emergency vehicle
                self.spawn_emergency_vehicle()
            elif event.key == pygame.K_2:  # NEW - Spawn car burst
                for _ in range(5):
                    self.spawn_car()
            elif event.key == pygame.K_ESCAPE:
                self.running = False
    
    def reset_simulation(self):
        """Reset the simulation"""
        self.cars.clear()
        self.total_cars_spawned = 0
        self.emergency_vehicles = 0
        
        # Reset intersections
        for intersection in self.intersections:
            intersection.rl_agent = RLAgent()
            intersection.light_state = 'red'
            intersection.light_timer = 0
            intersection.total_cars_passed = 0
            intersection.light_changes = 0
            intersection.performance_score = 0
    
def run(self):
    """Main simulation loop"""
    spawn_timer = 0
    data_save_timer = 0
    
    while self.running:
        self.handle_events()
        
        if not self.paused:
            # Spawn cars
            spawn_timer += 1
            if spawn_timer >= max(20, 60 - len(self.cars)):
                self.spawn_car()
                spawn_timer = 0
            
            # Update simulation
            self.update_simulation()
            
            # Save data periodically
            data_save_timer += 1
            if data_save_timer >= 3600:  # Every minute at 60 FPS
                self.data_logger.save_data()
                data_save_timer = 0
            
            self.frame_count += 1
        
        # Draw everything
        self.draw_gradient_background()
        self.draw_modern_roads()
        
        # Draw intersections
        for intersection in self.intersections:
            intersection.draw(self.screen, self.font)
        
        # Draw cars
        for car in self.cars:
            car.draw(self.screen)
        
        # Draw UI
        self.draw_modern_ui()
        self.draw_live_graphs()  # ADD THIS LINE
        
        pygame.display.flip()
        self.clock.tick(self.simulation_speed)
    
    # Save final data
    self.data_logger.save_data()
    pygame.quit()


class DataLogger:
    """Enhanced data logging and analytics"""
    def __init__(self):
        self.car_data = []
        self.intersection_data = []
        self.system_metrics = []
        self.start_time = datetime.now()
        
    def log_car(self, car):
        """Log car completion data"""
        travel_time = time.time() - car.spawn_time
        self.car_data.append({
            'car_id': id(car),
            'car_type': car.car_type,
            'spawn_time': car.spawn_time,
            'completion_time': time.time(),
            'travel_time': travel_time,
            'total_wait_time': car.total_wait_time,
            'start_x': car.start_x,
            'start_y': car.start_y,
            'target_intersection': car.target_intersection.id
        })
    
    def log_intersection_state(self, intersection):
        """Log intersection performance data"""
        self.intersection_data.append({
            'intersection_id': intersection.id,
            'timestamp': time.time(),
            'light_state': intersection.light_state,
            'vehicle_count': intersection.vehicle_count,
            'avg_wait_time': intersection.avg_wait_time,
            'total_cars_passed': intersection.total_cars_passed,
            'performance_score': intersection.performance_score,
            'epsilon': intersection.rl_agent.epsilon,
            'q_table_size': len(intersection.rl_agent.q_table)
        })
    
def save_data(self):
    """Save all collected data to Excel file with error handling"""
    try:
        # Create data directory if it doesn't exist
        os.makedirs("simulation_data", exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"simulation_data/traffic_simulation_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # Car data sheet
        if self.car_data:
            ws_cars = wb.active
            ws_cars.title = "Car Data"
            
            # Headers
            headers = list(self.car_data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws_cars.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Data
            for row, car_record in enumerate(self.car_data, 2):
                for col, value in enumerate(car_record.values(), 1):
                    # Handle datetime objects
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    ws_cars.cell(row=row, column=col, value=value)
        else:
            # Create empty sheet with headers
            ws_cars = wb.active
            ws_cars.title = "Car Data"
            headers = ['car_id', 'car_type', 'spawn_time', 'completion_time', 
                      'travel_time', 'total_wait_time', 'start_x', 'start_y', 'target_intersection']
            for col, header in enumerate(headers, 1):
                ws_cars.cell(row=1, column=col, value=header)
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        summary_data = [
            ["Simulation Start", self.start_time.isoformat()],
            ["Total Cars Processed", len(self.car_data)],
            ["Average Travel Time", np.mean([car['travel_time'] for car in self.car_data]) if self.car_data else 0],
            ["Average Wait Time", np.mean([car['total_wait_time'] for car in self.car_data]) if self.car_data else 0],
            ["Emergency Vehicles", len([car for car in self.car_data if car['car_type'] == 'emergency'])],
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row, column=1, value=label)
            ws_summary.cell(row=row, column=2, value=value)
        
        wb.save(filename)
        print(f"✅ Data successfully saved to {filename}")
        return True
        
    except Exception as e:
        print(f"❌ Error saving data: {e}")
        # Try saving to current directory as fallback
        try:
            fallback_filename = f"traffic_data_backup_{timestamp}.xlsx"
            wb.save(fallback_filename)
            print(f"✅ Backup saved to {fallback_filename}")
            return True
        except:
            print("❌ Could not save backup file")
            return False


# Initialize and run simulation
if __name__ == "__main__":
    # Add performance history to TrafficSimulation
    TrafficSimulation.performance_history = deque(maxlen=100)
    
    simulation = TrafficSimulation()
    simulation.run()