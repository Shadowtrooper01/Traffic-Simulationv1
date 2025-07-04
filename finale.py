import pygame
import numpy as np
import pandas as pd
import random
import math
import time
from datetime import datetime, timedelta
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
from sklearn.preprocessing import LabelEncoder, StandardScaler
import threading
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from collections import deque, defaultdict

# Initialize Pygame
pygame.init()

# Constants
WINDOW_WIDTH = 1600
WINDOW_HEIGHT = 1000
GRID_SIZE = 6
CELL_SIZE = 100
GRID_OFFSET_X = 50
GRID_OFFSET_Y = 50
ROAD_WIDTH = 8

# Enhanced Colors with gradients
BLACK = (0, 0, 0)
WHITE = (255, 255, 255)
RED = (255, 0, 0)
GREEN = (0, 255, 0)
YELLOW = (255, 255, 0)
BLUE = (0, 0, 255)
GRAY = (128, 128, 128)
LIGHT_GRAY = (200, 200, 200)
DARK_GRAY = (64, 64, 64)
ORANGE = (255, 165, 0)
PURPLE = (128, 0, 128)
CYAN = (0, 255, 255)
BROWN = (139, 69, 19)
DEEP_BLUE = (25, 25, 112)
NEON_GREEN = (57, 255, 20)
NEON_BLUE = (31, 81, 255)
DARK_PURPLE = (48, 25, 52)
GOLD = (255, 215, 0)

# Gradient colors for modern UI
GRADIENT_DARK = (15, 15, 25)
GRADIENT_LIGHT = (35, 35, 55)
ACCENT_BLUE = (64, 156, 255)
ACCENT_GREEN = (0, 230, 118)
ACCENT_RED = (255, 69, 58)
ACCENT_ORANGE = (255, 159, 10)
ACCENT_PURPLE = (179,158,151)

class RLAgent:
    """Deep Q-Network agent for traffic light optimization"""
    def __init__(self, state_size=8, action_size=4, learning_rate=0.001):
        self.state_size = state_size
        self.action_size = action_size
        self.memory = deque(maxlen=10000)
        self.epsilon = 1.0  # exploration rate
        self.epsilon_min = 0.01
        self.epsilon_decay = 0.995
        self.learning_rate = learning_rate
        self.q_table = defaultdict(lambda: np.zeros(action_size))
        self.performance_history = deque(maxlen=1000)
        self.reward_history = deque(maxlen=100)
        self.learning_enabled = True
        
    def get_state(self, intersection, nearby_cars):
        """Convert intersection state to feature vector"""
        waiting_cars = len([car for car in nearby_cars if car.waiting])
        emergency_cars = len([car for car in nearby_cars if car.car_type == 'emergency'])
        avg_wait_time = np.mean([car.wait_time for car in nearby_cars if car.waiting]) if waiting_cars > 0 else 0
        congestion_level = min(10, waiting_cars)
        
        # Light state encoding
        light_state_encoding = {'red': 0, 'yellow': 1, 'green': 2}
        light_state = light_state_encoding[intersection.light_state]
        
        # Time features
        time_in_state = intersection.light_timer / 60.0  # Normalize
        
        # Traffic flow features
        cars_approaching = len([car for car in nearby_cars if not car.waiting and 
                              math.sqrt((car.x - intersection.x)**2 + (car.y - intersection.y)**2) < 80])
        
        # Performance metrics
        efficiency_score = intersection.total_cars_passed / max(1, intersection.light_changes)
        
        state = np.array([
            waiting_cars / 10.0,  # Normalize to 0-1
            emergency_cars,
            avg_wait_time / 100.0,  # Normalize
            congestion_level / 10.0,
            light_state / 2.0,
            time_in_state,
            cars_approaching / 10.0,
            efficiency_score / 10.0
        ])
        
        return tuple(state.round(2))  # Round for discrete state space
    
    def get_action(self, state):
        """Choose action using epsilon-greedy policy"""
        if not self.learning_enabled or np.random.rand() <= self.epsilon:
            return random.randrange(self.action_size)
        
        q_values = self.q_table[state]
        return np.argmax(q_values)
    
    def remember(self, state, action, reward, next_state, done):
        """Store experience in replay memory"""
        self.memory.append((state, action, reward, next_state, done))
    
    def calculate_reward(self, intersection, action, prev_state):
        """Calculate reward for the action taken"""
        reward = 0
        
        # Penalize excessive waiting
        total_wait_time = sum(car.wait_time for car in intersection.cars_waiting)
        reward -= total_wait_time * 0.1
        
        # Reward for handling emergency vehicles
        if intersection.emergency_present:
            if action == 2:  # Green light for emergency
                reward += 50
            else:
                reward -= 20
        
        # Reward for efficient traffic flow
        if intersection.vehicle_count > 0 and intersection.light_state == 'green':
            reward += min(10, intersection.vehicle_count * 2)
        
        # Penalize unnecessary light changes
        if action == 1:  # Force change
            reward -= 5
        
        # Reward for maintaining good flow
        if intersection.total_cars_passed > 0:
            efficiency = intersection.total_cars_passed / max(1, intersection.light_changes)
            reward += efficiency * 2
        
        # Penalize high congestion
        if intersection.vehicle_count > 8:
            reward -= 15
        
        return reward
    
    def replay(self, batch_size=32):
        """Train the model on a batch of experiences"""
        if len(self.memory) < batch_size:
            return
        
        batch = random.sample(self.memory, batch_size)
        
        for state, action, reward, next_state, done in batch:
            target = reward
            if not done:
                target += 0.95 * np.max(self.q_table[next_state])  # Gamma = 0.95
            
            # Q-learning update
            self.q_table[state][action] += self.learning_rate * (target - self.q_table[state][action])
        
        # Decay epsilon
        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay
        
        self.reward_history.append(np.mean([exp[2] for exp in batch]))

class Car:
    def __init__(self, x, y, direction, speed, target_intersection, car_type='normal'):
        self.start_x = x
        self.start_y = y
        self.x = x
        self.y = y
        self.direction = direction
        self.base_speed = speed
        self.speed = speed
        self.target_intersection = target_intersection
        self.car_type = car_type
        self.color = self.get_car_color()
        self.size = 16 if car_type == 'emergency' else 12
        self.waiting = False
        self.wait_time = 0
        self.total_wait_time = 0
        self.spawn_time = time.time()
        self.completed = False
        self.path = []
        self.current_path_index = 0
        self.glow_effect = 0
        
    def get_car_color(self):
        if self.car_type == 'emergency':
            return ACCENT_RED
        colors = [NEON_BLUE, PURPLE, CYAN, ACCENT_ORANGE, ACCENT_GREEN, GOLD]
        return random.choice(colors)
        
    def update(self, intersections):
        should_stop = False
        stop_distance = 35
        
        for intersection in intersections:
            distance = math.sqrt((self.x - intersection.x)**2 + (self.y - intersection.y)**2)
            
            if distance < stop_distance and intersection.light_state == 'red':
                if self.car_type != 'emergency':
                    approaching = False
                    if self.direction == 0 and self.y > intersection.y:
                        approaching = True
                    elif self.direction == 1 and self.x < intersection.x:
                        approaching = True
                    elif self.direction == 2 and self.y < intersection.y:
                        approaching = True
                    elif self.direction == 3 and self.x > intersection.x:
                        approaching = True
                    
                    if approaching:
                        should_stop = True
                        break
        
        if should_stop:
            self.waiting = True
            self.wait_time += 1
            self.total_wait_time += 1
            self.speed = 0
            self.glow_effect = (self.glow_effect + 0.2) % (2 * math.pi)
        else:
            self.waiting = False
            if self.wait_time > 0:
                self.wait_time = 0
            self.speed = self.base_speed
            self.glow_effect = 0
            
            # Move car
            if self.direction == 0:
                self.y -= self.speed
            elif self.direction == 1:
                self.x += self.speed
            elif self.direction == 2:
                self.y += self.speed
            elif self.direction == 3:
                self.x -= self.speed
            
    def draw(self, screen):
        # Glow effect for waiting cars
        if self.waiting and self.glow_effect > 0:
            glow_radius = int(self.size + 4 + 2 * math.sin(self.glow_effect))
            glow_color = (*self.color[:3], 50)  # Semi-transparent
            # Create glow surface
            glow_surf = pygame.Surface((glow_radius * 2, glow_radius * 2), pygame.SRCALPHA)
            pygame.draw.circle(glow_surf, (*self.color, 30), (glow_radius, glow_radius), glow_radius)
            screen.blit(glow_surf, (self.x - glow_radius, self.y - glow_radius))
        
        # Main car body with rounded corners
        car_rect = pygame.Rect(self.x - self.size//2, self.y - self.size//2, self.size, self.size)
        pygame.draw.rect(screen, self.color, car_rect, border_radius=5)
        
        # Car outline
        pygame.draw.rect(screen, WHITE, car_rect, 2, border_radius=5)
        
        # Emergency vehicle lights
        if self.car_type == 'emergency':
            if (time.time() * 6) % 1 < 0.5:
                pygame.draw.circle(screen, WHITE, (int(self.x - 4), int(self.y - 4)), 3)
                pygame.draw.circle(screen, ACCENT_RED, (int(self.x + 4), int(self.y + 4)), 3)
        
        # Direction indicator with modern arrow
        arrow_size = 8
        arrow_color = WHITE if not self.waiting else YELLOW
        
        if self.direction == 0:  # North
            points = [(self.x, self.y - arrow_size), 
                     (self.x - 5, self.y + 3), 
                     (self.x + 5, self.y + 3)]
        elif self.direction == 1:  # East
            points = [(self.x + arrow_size, self.y), 
                     (self.x - 3, self.y - 5), 
                     (self.x - 3, self.y + 5)]
        elif self.direction == 2:  # South
            points = [(self.x, self.y + arrow_size), 
                     (self.x - 5, self.y - 3), 
                     (self.x + 5, self.y - 3)]
        else:  # West
            points = [(self.x - arrow_size, self.y), 
                     (self.x + 3, self.y - 5), 
                     (self.x + 3, self.y + 5)]
        
        pygame.draw.polygon(screen, arrow_color, points)

class Intersection:
    def __init__(self, row, col, x, y):
        self.row = row
        self.col = col
        self.x = x
        self.y = y
        self.id = f"{chr(65 + row)}{col + 1}"
        self.light_state = 'red'
        self.light_timer = 0
        self.light_duration = 180
        self.cars_waiting = []
        self.congestion_level = 0
        self.vehicle_count = 0
        self.emergency_present = False
        self.total_cars_passed = 0
        self.avg_wait_time = 0
        self.light_changes = 0
        self.rl_agent = RLAgent()
        self.previous_state = None
        self.last_action = 0
        self.performance_score = 0
        self.animation_offset = 0
        
    def get_nearby_cars(self, cars):
        nearby = []
        for car in cars:
            distance = math.sqrt((car.x - self.x)**2 + (car.y - self.y)**2)
            if distance < 60:
                nearby.append(car)
        return nearby
        
    def update(self, cars, learning_enabled=True):
        self.light_timer += 1
        self.animation_offset = (self.animation_offset + 0.1) % (2 * math.pi)
        
        # Get nearby cars and update metrics
        nearby_cars = self.get_nearby_cars(cars)
        waiting_cars = [car for car in nearby_cars if car.waiting]
        
        self.cars_waiting = waiting_cars
        self.vehicle_count = len(waiting_cars)
        self.emergency_present = any(car.car_type == 'emergency' for car in nearby_cars)
        self.congestion_level = min(10, len(waiting_cars))
        
        if waiting_cars:
            self.avg_wait_time = sum(car.wait_time for car in waiting_cars) / len(waiting_cars)
        
        # RL-based traffic light control
        current_state = self.rl_agent.get_state(self, nearby_cars)
        
        if self.previous_state is not None and learning_enabled:
            # Calculate reward and learn
            reward = self.rl_agent.calculate_reward(self, self.last_action, self.previous_state)
            self.rl_agent.remember(self.previous_state, self.last_action, reward, current_state, False)
            
            # Train the agent
            if len(self.rl_agent.memory) > 100:
                self.rl_agent.replay(32)
        
        # Get action from RL agent
        action = self.rl_agent.get_action(current_state)
        self.last_action = action
        self.previous_state = current_state
        
        # Execute action
        self.execute_action(action)
        
        # Update performance score
        self.update_performance_score()
        
    def execute_action(self, action):
        """Execute the action chosen by RL agent"""
        # Actions: 0=maintain, 1=force_change, 2=extend_green, 3=quick_change
        
        if action == 1:  # Force change
            self.force_light_change()
        elif action == 2 and self.light_state == 'green':  # Extend green
            self.light_duration = min(self.light_duration + 60, 400)
        elif action == 3:  # Quick change
            self.light_timer = max(self.light_timer, self.light_duration - 30)
        
        # Normal light cycle with dynamic timing
        base_durations = self.get_dynamic_durations()
        
        if self.light_timer >= self.light_duration:
            if self.light_state == 'red':
                self.light_state = 'green'
                self.light_duration = base_durations['green']
                self.light_changes += 1
            elif self.light_state == 'green':
                self.light_state = 'yellow'
                self.light_duration = base_durations['yellow']
                self.light_changes += 1
            else:  # yellow
                self.light_state = 'red'
                self.light_duration = base_durations['red']
                self.light_changes += 1
            self.light_timer = 0
    
    
    def get_dynamic_durations(self):
        """Calculate dynamic light durations based on conditions"""
        base_green = 240
        base_red = 180
        base_yellow = 60
        
        # Emergency override
        if self.emergency_present:
            return {'green': 360, 'red': 60, 'yellow': 40}
        
        # Adjust based on congestion
        if self.vehicle_count > 8:
            base_green = 420
        elif self.vehicle_count > 5:
            base_green = 300
        elif self.vehicle_count == 0:
            base_green = 120
            base_red = 300
        
        return {'green': base_green, 'red': base_red, 'yellow': base_yellow}
    
    def force_light_change(self):
        """Force immediate light change"""
        if self.light_state == 'red':
            self.light_state = 'green'
            self.light_duration = self.get_dynamic_durations()['green']
        elif self.light_state == 'green':
            self.light_state = 'yellow'
            self.light_duration = self.get_dynamic_durations()['yellow']
        else:
            self.light_state = 'red' 
            self.light_duration = self.get_dynamic_durations()['red']
        
        self.light_timer = 0
        self.light_changes += 1
    
    def update_performance_score(self):
        """Calculate intersection performance score"""
        efficiency = self.total_cars_passed / max(1, self.light_changes)
        wait_penalty = self.avg_wait_time / 100.0
        congestion_penalty = self.vehicle_count / 10.0
        
        self.performance_score = max(0, efficiency - wait_penalty - congestion_penalty)
        
    def draw(self, screen, font):
        # Animated intersection base with gradient effect
        base_radius = 22
        glow_radius = int(base_radius + 3 * math.sin(self.animation_offset))
        
        # Glow effect
        glow_surf = pygame.Surface((glow_radius * 2, glow_radius * 2), pygame.SRCALPHA)
        pygame.draw.circle(glow_surf, (*DARK_GRAY, 100), (glow_radius, glow_radius), glow_radius)
        screen.blit(glow_surf, (self.x - glow_radius, self.y - glow_radius))
        
        # Main intersection
        pygame.draw.circle(screen, GRADIENT_LIGHT, (int(self.x), int(self.y)), base_radius)
        pygame.draw.circle(screen, LIGHT_GRAY, (int(self.x), int(self.y)), base_radius, 2)
        
        # Traffic light with modern design
        light_colors = {'red': ACCENT_RED, 'yellow': ACCENT_ORANGE, 'green': ACCENT_GREEN}
        light_color = light_colors[self.light_state]
        light_radius = 14
        
        # Light glow
        light_glow_surf = pygame.Surface((light_radius * 3, light_radius * 3), pygame.SRCALPHA)
        pygame.draw.circle(light_glow_surf, (*light_color, 80), (light_radius * 1.5, light_radius * 1.5), light_radius * 1.5)
        screen.blit(light_glow_surf, (self.x - light_radius * 1.5, self.y - light_radius * 1.5))
        
        # Main light
        pygame.draw.circle(screen, light_color, (int(self.x), int(self.y)), light_radius)
        pygame.draw.circle(screen, WHITE, (int(self.x), int(self.y)), light_radius, 3)
        
        # Intersection ID with modern styling
        id_surface = font.render(self.id, True, WHITE)
        id_rect = id_surface.get_rect(center=(self.x, self.y - 45))
        
        # ID background
        bg_rect = id_rect.inflate(12, 6)
        pygame.draw.rect(screen, GRADIENT_DARK, bg_rect, border_radius=8)
        pygame.draw.rect(screen, ACCENT_BLUE, bg_rect, 2, border_radius=8)
        screen.blit(id_surface, id_rect)
        
        # Performance indicator
        if self.vehicle_count > 0:
            perf_color = ACCENT_GREEN if self.performance_score > 5 else ACCENT_ORANGE if self.performance_score > 2 else ACCENT_RED
            perf_radius = int(12 + self.performance_score * 2)
            
            pygame.draw.circle(screen, perf_color, (int(self.x + 30), int(self.y - 30)), perf_radius, 4)
            
            # Vehicle count
            count_text = font.render(str(self.vehicle_count), True, WHITE)
            screen.blit(count_text, (self.x + 25, self.y - 35))

class TrafficSimulation:
    def __init__(self):
        self.screen = pygame.display.set_mode((WINDOW_WIDTH, WINDOW_HEIGHT))
        pygame.display.set_caption("AI Traffic Management System - Reinforcement Learning")
        self.clock = pygame.time.Clock()
        self.font = pygame.font.Font(None, 18)
        self.big_font = pygame.font.Font(None, 28)
        self.title_font = pygame.font.Font(None, 36)
        
        # Create intersections
        self.intersections = []
        for row in range(GRID_SIZE):
            for col in range(GRID_SIZE):
                x = GRID_OFFSET_X + col * CELL_SIZE + CELL_SIZE // 2
                y = GRID_OFFSET_Y + row * CELL_SIZE + CELL_SIZE // 2
                intersection = Intersection(row, col, x, y)
                self.intersections.append(intersection)
        
        self.cars = []
        self.data_logger = DataLogger()
        self.running = True
        self.paused = False
        self.simulation_speed = 60  # FPS
        self.speed_multiplier = 1.0
        self.frame_count = 0
        self.learning_enabled = True
        
        # Enhanced statistics
        self.total_cars_spawned = 0
        self.total_wait_time = 0
        self.emergency_vehicles = 0
        self.system_efficiency = 0
        self.avg_performance = 0
        self.learning_progress = 0
        # Performance tracking
        self.performance_history = deque(maxlen=100)
        self.efficiency_history = deque(maxlen=100)
        self.learning_history = deque(maxlen=100)
        self.reward_history = deque(maxlen=100)
        
        # UI state
        self.show_rl_info = True
        self.show_performance = True
        self.show_live_graphs = True
        
    def spawn_car(self):
        if len(self.cars) > 50:
            return
        
        # Enhanced spawn logic with better distribution
        spawn_points = []
        
        # Edge spawns with better positioning
        for i in range(GRID_SIZE):
            x = GRID_OFFSET_X + i * CELL_SIZE + CELL_SIZE // 2
            y = GRID_OFFSET_Y - 30
            spawn_points.append((x, y, 2))
            
            x = GRID_OFFSET_X + i * CELL_SIZE + CELL_SIZE // 2
            y = GRID_OFFSET_Y + GRID_SIZE * CELL_SIZE + 30
            spawn_points.append((x, y, 0))
            
            x = GRID_OFFSET_X - 30
            y = GRID_OFFSET_Y + i * CELL_SIZE + CELL_SIZE // 2
            spawn_points.append((x, y, 1))
            
            x = GRID_OFFSET_X + GRID_SIZE * CELL_SIZE + 30
            y = GRID_OFFSET_Y + i * CELL_SIZE + CELL_SIZE // 2
            spawn_points.append((x, y, 3))
        
        x, y, direction = random.choice(spawn_points)
        
        # Dynamic car type distribution
        car_type = 'emergency' if random.random() < 0.06 else 'normal'
        speed = random.uniform(1.8, 3.0) if car_type == 'normal' else random.uniform(3.0, 4.0)
        speed *= self.speed_multiplier
        
        target = random.choice(self.intersections)
        car = Car(x, y, direction, speed, target, car_type)
        self.cars.append(car)
        self.total_cars_spawned += 1
        
        if car_type == 'emergency':
            self.emergency_vehicles += 1
    
    def update_simulation(self):
        # Update intersections with RL
        for intersection in self.intersections:
            intersection.update(self.cars, self.learning_enabled)
        
        # Update cars
        cars_to_remove = []
        for car in self.cars:
            car.update(self.intersections)
            
            if (car.x < -100 or car.x > WINDOW_WIDTH + 100 or 
                car.y < -100 or car.y > WINDOW_HEIGHT + 100):
                car.completed = True
                self.data_logger.log_car(car)
                cars_to_remove.append(car)
        
        for car in cars_to_remove:
            self.cars.remove(car)
        
        # Update system metrics
        self.update_system_metrics()
    
    def update_system_metrics(self):
        """Update system-wide performance metrics"""
        if self.intersections:
            self.avg_performance = np.mean([i.performance_score for i in self.intersections])
            self.performance_history.append(self.avg_performance)  # ADD THIS LINE
            
            total_cars = sum(len(i.cars_waiting) for i in self.intersections)
            total_wait = sum(sum(car.wait_time for car in i.cars_waiting) for i in self.intersections)
            self.system_efficiency = max(0, 100 - (total_wait / max(1, total_cars)))
            self.efficiency_history.append(self.system_efficiency)  # ADD THIS LINE
            
            # Learning progress based on epsilon decay
            avg_epsilon = np.mean([i.rl_agent.epsilon for i in self.intersections])
            self.learning_progress = (1 - avg_epsilon) * 100
            self.learning_history.append(self.learning_progress)  # ADD THIS LINE
            
            # Average reward tracking
            avg_reward = 0
            reward_count = 0
            for intersection in self.intersections:
                if intersection.rl_agent.reward_history:
                    avg_reward += np.mean(list(intersection.rl_agent.reward_history))
                    reward_count += 1
            if reward_count > 0:
                avg_reward /= reward_count
                self.reward_history.append(avg_reward)  # ADD THIS LINE
        
    def draw_gradient_background(self):
        """Draw modern gradient background"""
        for y in range(WINDOW_HEIGHT):
            ratio = y / WINDOW_HEIGHT


class DataLogger:
    """Enhanced data logging and analytics"""
    def __init__(self):
        self.car_data = []
        self.intersection_data = []
        self.system_metrics = []
        self.start_time = datetime.now()
        
    def log_car(self, car):
        """Log car completion data"""
        travel_time = time.time() - car.spawn_time
        self.car_data.append({
            'car_id': id(car),
            'car_type': car.car_type,
            'spawn_time': car.spawn_time,
            'completion_time': time.time(),
            'travel_time': travel_time,
            'total_wait_time': car.total_wait_time,
            'start_x': car.start_x,
            'start_y': car.start_y,
            'target_intersection': car.target_intersection.id
        })
    
    def log_intersection_state(self, intersection):
        """Log intersection performance data"""
        self.intersection_data.append({
            'intersection_id': intersection.id,
            'timestamp': time.time(),
            'light_state': intersection.light_state,
            'vehicle_count': intersection.vehicle_count,
            'avg_wait_time': intersection.avg_wait_time,
            'total_cars_passed': intersection.total_cars_passed,
            'performance_score': intersection.performance_score,
            'epsilon': intersection.rl_agent.epsilon,
            'q_table_size': len(intersection.rl_agent.q_table)
        })
    
def save_data(self):
    """Save all collected data to Excel file with error handling"""
    try:
        # Create data directory if it doesn't exist
        os.makedirs("simulation_data", exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"simulation_data/traffic_simulation_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # Car data sheet
        if self.car_data:
            ws_cars = wb.active
            ws_cars.title = "Car Data"
            
            # Headers
            headers = list(self.car_data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws_cars.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Data
            for row, car_record in enumerate(self.car_data, 2):
                for col, value in enumerate(car_record.values(), 1):
                    # Handle datetime objects
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    ws_cars.cell(row=row, column=col, value=value)
        else:
            # Create empty sheet with headers
            ws_cars = wb.active
            ws_cars.title = "Car Data"
            headers = ['car_id', 'car_type', 'spawn_time', 'completion_time', 
                      'travel_time', 'total_wait_time', 'start_x', 'start_y', 'target_intersection']
            for col, header in enumerate(headers, 1):
                ws_cars.cell(row=1, column=col, value=header)
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        summary_data = [
            ["Simulation Start", self.start_time.isoformat()],
            ["Total Cars Processed", len(self.car_data)],
            ["Average Travel Time", np.mean([car['travel_time'] for car in self.car_data]) if self.car_data else 0],
            ["Average Wait Time", np.mean([car['total_wait_time'] for car in self.car_data]) if self.car_data else 0],
            ["Emergency Vehicles", len([car for car in self.car_data if car['car_type'] == 'emergency'])],
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row, column=1, value=label)
            ws_summary.cell(row=row, column=2, value=value)
        
        wb.save(filename)
        print(f"✅ Data successfully saved to {filename}")
        return True
        
    except Exception as e:
        print(f"❌ Error saving data: {e}")
        # Try saving to current directory as fallback
        try:
            fallback_filename = f"traffic_data_backup_{timestamp}.xlsx"
            wb.save(fallback_filename)
            print(f"✅ Backup saved to {fallback_filename}")
            return True
        except:
            print("❌ Could not save backup file")
            return False


# Initialize and run simulation
if __name__ == "__main__":
    # Add performance history to TrafficSimulation
    TrafficSimulation.performance_history = deque(maxlen=100)
    
    simulation = TrafficSimulation()
    simulation.run()