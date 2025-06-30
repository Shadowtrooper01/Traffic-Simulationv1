import pygame
import numpy as np
import pandas as pd
import random
import math
import time
from datetime import datetime, timedelta
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
from sklearn.preprocessing import LabelEncoder, StandardScaler
import threading
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Initialize Pygame
pygame.init()

# Constants
WINDOW_WIDTH = 1400
WINDOW_HEIGHT = 900
GRID_SIZE = 6  # 6x6 grid for better visibility
CELL_SIZE = 100
GRID_OFFSET_X = 50
GRID_OFFSET_Y = 50
ROAD_WIDTH = 8

# Colors
BLACK = (0, 0, 0)
WHITE = (255, 255, 255)
RED = (255, 0, 0)
GREEN = (0, 255, 0)
YELLOW = (255, 255, 0)
BLUE = (0, 0, 255)
GRAY = (128, 128, 128)
LIGHT_GRAY = (200, 200, 200)
DARK_GRAY = (64, 64, 64)
ORANGE = (255, 165, 0)
PURPLE = (128, 0, 128)
CYAN = (0, 255, 255)
BROWN = (139, 69, 19)

# Traffic light states
LIGHT_COLORS = {
    'red': RED,
    'yellow': YELLOW,
    'green': GREEN
}

class Car:
    def __init__(self, x, y, direction, speed, target_intersection, car_type='normal'):
        self.start_x = x
        self.start_y = y
        self.x = x
        self.y = y
        self.direction = direction  # 0=North, 1=East, 2=South, 3=West
        self.base_speed = speed
        self.speed = speed
        self.target_intersection = target_intersection
        self.car_type = car_type  # 'normal', 'emergency'
        self.color = RED if car_type == 'emergency' else random.choice([BLUE, PURPLE, CYAN, ORANGE, BROWN])
        self.size = 14 if car_type == 'emergency' else 10
        self.waiting = False
        self.wait_time = 0
        self.total_wait_time = 0
        self.spawn_time = time.time()
        self.completed = False
        self.path = []
        self.current_path_index = 0
        
    def update(self, intersections):
        # Check if car should stop at red light
        should_stop = False
        stop_distance = 30
        
        for intersection in intersections:
            # Calculate distance to intersection
            distance = math.sqrt((self.x - intersection.x)**2 + (self.y - intersection.y)**2)
            
            # Check if approaching intersection with red light
            if distance < stop_distance and intersection.light_state == 'red':
                # Emergency vehicles don't stop
                if self.car_type != 'emergency':
                    # Check if car is approaching from correct direction
                    approaching = False
                    if self.direction == 0 and self.y > intersection.y:  # North, car is south of intersection
                        approaching = True
                    elif self.direction == 1 and self.x < intersection.x:  # East, car is west of intersection
                        approaching = True
                    elif self.direction == 2 and self.y < intersection.y:  # South, car is north of intersection
                        approaching = True
                    elif self.direction == 3 and self.x > intersection.x:  # West, car is east of intersection
                        approaching = True
                    
                    if approaching:
                        should_stop = True
                        break
        
        if should_stop:
            self.waiting = True
            self.wait_time += 1
            self.total_wait_time += 1
            self.speed = 0
        else:
            self.waiting = False
            if self.wait_time > 0:
                self.wait_time = 0
            self.speed = self.base_speed
            
            # Move car
            if self.direction == 0:  # North
                self.y -= self.speed
            elif self.direction == 1:  # East
                self.x += self.speed
            elif self.direction == 2:  # South
                self.y += self.speed
            elif self.direction == 3:  # West
                self.x -= self.speed
            
    def draw(self, screen):
        # Draw car as rectangle with rounded corners
        car_rect = pygame.Rect(self.x - self.size//2, self.y - self.size//2, self.size, self.size)
        pygame.draw.rect(screen, self.color, car_rect, border_radius=3)
        
        # Draw emergency vehicle lights
        if self.car_type == 'emergency':
            # Flashing effect
            if (time.time() * 4) % 1 < 0.5:
                pygame.draw.circle(screen, WHITE, (int(self.x - 3), int(self.y - 3)), 2)
                pygame.draw.circle(screen, WHITE, (int(self.x + 3), int(self.y + 3)), 2)
        
        # Draw direction indicator
        arrow_size = 6
        if self.direction == 0:  # North - triangle pointing up
            points = [(self.x, self.y - arrow_size), 
                     (self.x - 4, self.y + 2), 
                     (self.x + 4, self.y + 2)]
        elif self.direction == 1:  # East - triangle pointing right
            points = [(self.x + arrow_size, self.y), 
                     (self.x - 2, self.y - 4), 
                     (self.x - 2, self.y + 4)]
        elif self.direction == 2:  # South - triangle pointing down
            points = [(self.x, self.y + arrow_size), 
                     (self.x - 4, self.y - 2), 
                     (self.x + 4, self.y - 2)]
        else:  # West - triangle pointing left
            points = [(self.x - arrow_size, self.y), 
                     (self.x + 2, self.y - 4), 
                     (self.x + 2, self.y + 4)]
        
        pygame.draw.polygon(screen, WHITE, points)

class Intersection:
    def __init__(self, row, col, x, y):
        self.row = row
        self.col = col
        self.x = x
        self.y = y
        self.id = f"{chr(65 + row)}{col + 1}"
        self.light_state = 'red'
        self.light_timer = 0
        self.light_duration = 180  # frames (~3 seconds at 60 FPS)
        self.cars_waiting = []
        self.congestion_level = 0
        self.vehicle_count = 0
        self.emergency_present = False
        self.priority_score = 0
        self.action_recommendation = "Normal_Schedule"
        self.total_cars_passed = 0
        self.avg_wait_time = 0
        self.light_changes = 0
        
    def update(self, cars):
        self.light_timer += 1
        
        # Count nearby cars and detect emergency vehicles
        nearby_cars = []
        waiting_cars = []
        emergency_nearby = False
        
        for car in cars:
            distance = math.sqrt((car.x - self.x)**2 + (car.y - self.y)**2)
            if distance < 50:  # Detection radius
                nearby_cars.append(car)
                if car.waiting and distance < 35:
                    waiting_cars.append(car)
                if car.car_type == 'emergency':
                    emergency_nearby = True
        
        self.cars_waiting = waiting_cars
        self.vehicle_count = len(waiting_cars)
        self.emergency_present = emergency_nearby
        self.congestion_level = min(10, len(waiting_cars))
        
        # Calculate average wait time for waiting cars
        if waiting_cars:
            self.avg_wait_time = sum(car.wait_time for car in waiting_cars) / len(waiting_cars)
        
        # Individual traffic light logic
        green_duration = 240  # Base green time
        red_duration = 180    # Base red time
        yellow_duration = 60  # Yellow time
        
        # Adjust timing based on conditions
        if self.emergency_present:
            # Emergency override - turn green immediately
            if self.light_state == 'red':
                self.light_state = 'green'
                self.light_timer = 0
                self.light_changes += 1
            green_duration = 360  # Longer green for emergency
        elif self.vehicle_count > 5:
            # Heavy congestion - longer green
            green_duration = 300
        elif self.vehicle_count > 8:
            # Very heavy congestion - much longer green
            green_duration = 420
        elif self.vehicle_count == 0:
            # No cars waiting - shorter green, longer red
            green_duration = 120
            red_duration = 300
        
        # Normal traffic light cycle
        if self.light_timer >= self.light_duration:
            if self.light_state == 'red':
                self.light_state = 'green'
                self.light_duration = green_duration
                self.light_changes += 1
            elif self.light_state == 'green':
                self.light_state = 'yellow'
                self.light_duration = yellow_duration
                self.light_changes += 1
            else:  # yellow
                self.light_state = 'red'
                self.light_duration = red_duration
                self.light_changes += 1
            self.light_timer = 0
            
    def draw(self, screen, font):
        # Draw intersection base
        pygame.draw.circle(screen, DARK_GRAY, (int(self.x), int(self.y)), 18)
        
        # Draw traffic light with better visibility
        light_color = LIGHT_COLORS[self.light_state]
        pygame.draw.circle(screen, light_color, (int(self.x), int(self.y)), 12)
        pygame.draw.circle(screen, BLACK, (int(self.x), int(self.y)), 12, 2)
        
        # Draw intersection ID
        text = font.render(self.id, True, WHITE)
        text_rect = text.get_rect(center=(self.x, self.y - 40))
        screen.blit(text, text_rect)
        
        # Draw congestion level indicator
        if self.vehicle_count > 0:
            congestion_color = RED if self.vehicle_count > 6 else ORANGE if self.vehicle_count > 3 else YELLOW
            congestion_radius = min(25, 8 + self.vehicle_count * 2)
            pygame.draw.circle(screen, congestion_color, (int(self.x + 25), int(self.y - 25)), 
                             congestion_radius, 3)
            
            # Show vehicle count
            count_text = font.render(str(self.vehicle_count), True, WHITE)
            screen.blit(count_text, (self.x + 20, self.y - 30))

class TrafficModel:
    def __init__(self):
        self.le_intersection = LabelEncoder()
        self.le_day = LabelEncoder()
        self.scaler = StandardScaler()
        self.traffic_classifier = RandomForestClassifier(n_estimators=50, random_state=42)
        self.is_trained = False
        self.current_day = 'Monday'
        self.current_hour = 8
        
    def generate_training_data(self):
        """Generate training data for the model"""
        intersections = [f"{chr(65 + i)}{j + 1}" for i in range(GRID_SIZE) for j in range(GRID_SIZE)]
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        
        data = []
        for _ in range(800):  # Smaller dataset for faster training
            intersection = random.choice(intersections)
            day = random.choice(days)
            hour = random.randint(0, 23)
            
            # Simple congestion simulation
            base_congestion = 2.0
            if hour in [7, 8, 9, 17, 18, 19]:  # Rush hour
                base_congestion *= random.uniform(2, 3)
            if day in ['Saturday', 'Sunday']:
                base_congestion *= random.uniform(0.5, 1.5)
                
            congestion_time = max(0.5, base_congestion + np.random.normal(0, 0.5))
            vehicle_count = max(0, int(congestion_time * random.uniform(2, 5)))
            emergency_present = random.random() < 0.05
            
            # Determine action
            if emergency_present:
                action = 'Open_Immediately'
            elif congestion_time > 6:
                action = 'Open_Soon'
            elif congestion_time > 3:
                action = 'Normal_Schedule'
            else:
                action = 'Delay_Opening'
                
            data.append({
                'intersection': intersection,
                'day': day,
                'hour': hour,
                'congestion_time': congestion_time,
                'vehicle_count': vehicle_count,
                'emergency_present': emergency_present,
                'action': action
            })
        
        return pd.DataFrame(data)
    
    def train(self):
        """Train the traffic model"""
        df = self.generate_training_data()
        
        # Encode categorical variables
        df['intersection_encoded'] = self.le_intersection.fit_transform(df['intersection'])
        df['day_encoded'] = self.le_day.fit_transform(df['day'])
        
        # Features
        X = df[['intersection_encoded', 'day_encoded', 'hour', 
                'congestion_time', 'vehicle_count', 'emergency_present']]
        y = df['action']
        
        # Scale and train
        X_scaled = self.scaler.fit_transform(X)
        self.traffic_classifier.fit(X_scaled, y)
        self.is_trained = True
        
    def predict_action(self, intersection_id, congestion_time, vehicle_count, emergency_present):
        """Predict traffic light action"""
        if not self.is_trained:
            return "Normal_Schedule"
            
        try:
            # Prepare input
            input_data = np.array([[
                self.le_intersection.transform([intersection_id])[0],
                self.le_day.transform([self.current_day])[0],
                self.current_hour,
                congestion_time,
                vehicle_count,
                emergency_present
            ]])
            
            input_scaled = self.scaler.transform(input_data)
            prediction = self.traffic_classifier.predict(input_scaled)[0]
            return prediction
        except:
            return "Normal_Schedule"

class DataLogger:
    def __init__(self):
        self.session_data = []
        self.car_data = []
        self.intersection_data = []
        self.start_time = datetime.now()
        
    def log_car(self, car):
        """Log car data"""
        self.car_data.append({
            'car_id': id(car),
            'car_type': car.car_type,
            'spawn_time': car.spawn_time,
            'total_wait_time': car.total_wait_time,
            'completed': car.completed,
            'start_x': car.start_x,
            'start_y': car.start_y,
            'direction': car.direction,
            'target_intersection': car.target_intersection.id if car.target_intersection else 'None'
        })
    
    def log_intersection(self, intersection):
        """Log intersection data"""
        self.intersection_data.append({
            'intersection_id': intersection.id,
            'row': intersection.row,
            'col': intersection.col,
            'total_cars_passed': intersection.total_cars_passed,
            'light_changes': intersection.light_changes,
            'avg_wait_time': intersection.avg_wait_time,
            'final_vehicle_count': intersection.vehicle_count,
            'final_congestion_level': intersection.congestion_level
        })
    
    def save_to_excel(self, filename="traffic_simulation_data.xlsx"):
        """Save all data to Excel file"""
        try:
            wb = Workbook()
            
            # Session Summary Sheet
            ws1 = wb.active
            ws1.title = "Session Summary"
            
            # Headers with styling
            headers = ['Metric', 'Value']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Session data
            session_summary = [
                ['Session Start Time', self.start_time.strftime('%Y-%m-%d %H:%M:%S')],
                ['Session End Time', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Total Duration (minutes)', f"{(datetime.now() - self.start_time).total_seconds() / 60:.2f}"],
                ['Total Cars Spawned', len(self.car_data)],
                ['Emergency Vehicles', len([car for car in self.car_data if car['car_type'] == 'emergency'])],
                ['Average Wait Time per Car', f"{np.mean([car['total_wait_time'] for car in self.car_data]):.2f}" if self.car_data else "0"],
                ['Total Intersections', len(self.intersection_data)]
            ]
            
            for row, (metric, value) in enumerate(session_summary, 2):
                ws1.cell(row=row, column=1, value=metric)
                ws1.cell(row=row, column=2, value=value)
            
            # Car Data Sheet
            if self.car_data:
                ws2 = wb.create_sheet("Car Data")
                car_df = pd.DataFrame(self.car_data)
                
                # Headers
                for col, column in enumerate(car_df.columns, 1):
                    cell = ws2.cell(row=1, column=col, value=column)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                
                # Data
                for row, record in enumerate(car_df.to_dict('records'), 2):
                    for col, (key, value) in enumerate(record.items(), 1):
                        ws2.cell(row=row, column=col, value=value)
            
            # Intersection Data Sheet
            if self.intersection_data:
                ws3 = wb.create_sheet("Intersection Data")
                intersection_df = pd.DataFrame(self.intersection_data)
                
                # Headers
                for col, column in enumerate(intersection_df.columns, 1):
                    cell = ws3.cell(row=1, column=col, value=column)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                
                # Data
                for row, record in enumerate(intersection_df.to_dict('records'), 2):
                    for col, (key, value) in enumerate(record.items(), 1):
                        ws3.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            print(f"✅ Data saved to {filename}")
            return True
            
        except Exception as e:
            print(f"❌ Error saving Excel file: {e}")
            return False

class TrafficSimulation:
    def __init__(self):
        self.screen = pygame.display.set_mode((WINDOW_WIDTH, WINDOW_HEIGHT))
        pygame.display.set_caption("Smart Traffic Management System - Enhanced Simulation")
        self.clock = pygame.time.Clock()
        self.font = pygame.font.Font(None, 20)
        self.big_font = pygame.font.Font(None, 32)
        
        # Create grid of intersections
        self.intersections = []
        for row in range(GRID_SIZE):
            for col in range(GRID_SIZE):
                x = GRID_OFFSET_X + col * CELL_SIZE + CELL_SIZE // 2
                y = GRID_OFFSET_Y + row * CELL_SIZE + CELL_SIZE // 2
                intersection = Intersection(row, col, x, y)
                self.intersections.append(intersection)
        
        self.cars = []
        self.model = TrafficModel()
        self.data_logger = DataLogger()
        self.running = True
        self.paused = False
        self.simulation_speed = 1
        self.frame_count = 0
        
        # Statistics
        self.total_cars_spawned = 0
        self.total_wait_time = 0
        self.emergency_vehicles = 0
        
        # Start model training in background
        threading.Thread(target=self.model.train, daemon=True).start()
        
    def spawn_car(self):
        """Spawn a new car at road entrances"""
        if len(self.cars) > 40:  # Limit cars for performance
            return
            
        # Define spawn points at road entrances (aligned to roads)
        spawn_points = []
        
        # Top edge entrances
        for i in range(GRID_SIZE):
            x = GRID_OFFSET_X + i * CELL_SIZE + CELL_SIZE // 2
            y = GRID_OFFSET_Y - 20
            spawn_points.append((x, y, 2))  # Direction South
        
        # Bottom edge entrances
        for i in range(GRID_SIZE):
            x = GRID_OFFSET_X + i * CELL_SIZE + CELL_SIZE // 2
            y = GRID_OFFSET_Y + GRID_SIZE * CELL_SIZE + 20
            spawn_points.append((x, y, 0))  # Direction North
        
        # Left edge entrances
        for i in range(GRID_SIZE):
            x = GRID_OFFSET_X - 20
            y = GRID_OFFSET_Y + i * CELL_SIZE + CELL_SIZE // 2
            spawn_points.append((x, y, 1))  # Direction East
        
        # Right edge entrances
        for i in range(GRID_SIZE):
            x = GRID_OFFSET_X + GRID_SIZE * CELL_SIZE + 20
            y = GRID_OFFSET_Y + i * CELL_SIZE + CELL_SIZE // 2
            spawn_points.append((x, y, 3))  # Direction West
        
        # Choose random spawn point
        x, y, direction = random.choice(spawn_points)
        
        # Determine car type
        car_type = 'emergency' if random.random() < 0.08 else 'normal'
        speed = random.uniform(1.5, 2.5) if car_type == 'normal' else random.uniform(2.5, 3.5)
        
        # Random target intersection
        target = random.choice(self.intersections)
        
        car = Car(x, y, direction, speed, target, car_type)
        self.cars.append(car)
        self.total_cars_spawned += 1
        
        if car_type == 'emergency':
            self.emergency_vehicles += 1
    
    def update_intersections(self):
        """Update all intersections independently"""
        for intersection in self.intersections:
            intersection.update(self.cars)
            
            # Get AI recommendation if model is trained
            if self.model.is_trained:
                congestion_time = max(0.5, intersection.vehicle_count * 0.8 + 
                                    (3 if intersection.emergency_present else 0))
                
                recommendation = self.model.predict_action(
                    intersection.id, congestion_time, intersection.vehicle_count,
                    intersection.emergency_present
                )
                intersection.action_recommendation = recommendation
    
    def update_cars(self):
        """Update all cars"""
        cars_to_remove = []
        
        for car in self.cars:
            car.update(self.intersections)
            
            # Remove cars that are off screen
            if (car.x < -50 or car.x > WINDOW_WIDTH + 50 or 
                car.y < -50 or car.y > WINDOW_HEIGHT + 50):
                car.completed = True
                self.data_logger.log_car(car)
                cars_to_remove.append(car)
        
        for car in cars_to_remove:
            self.cars.remove(car)
    
    def draw_roads(self):
        """Draw the road network"""
        # Draw horizontal roads
        for i in range(GRID_SIZE):
            y = GRID_OFFSET_Y + i * CELL_SIZE + CELL_SIZE // 2
            # Road surface
            pygame.draw.line(self.screen, GRAY, 
                           (0, y), (WINDOW_WIDTH, y), ROAD_WIDTH * 2)
            # Road markings
            pygame.draw.line(self.screen, WHITE, 
                           (0, y), (WINDOW_WIDTH, y), 2)
        
        # Draw vertical roads
        for i in range(GRID_SIZE):
            x = GRID_OFFSET_X + i * CELL_SIZE + CELL_SIZE // 2
            # Road surface
            pygame.draw.line(self.screen, GRAY, 
                           (x, 0), (x, WINDOW_HEIGHT), ROAD_WIDTH * 2)
            # Road markings
            pygame.draw.line(self.screen, WHITE, 
                           (x, 0), (x, WINDOW_HEIGHT), 2)
    
    def draw_ui(self):
        """Draw user interface"""
        # Background panel
        ui_rect = pygame.Rect(GRID_OFFSET_X + GRID_SIZE * CELL_SIZE + 30, 50, 370, 800)
        pygame.draw.rect(self.screen, DARK_GRAY, ui_rect)
        pygame.draw.rect(self.screen, WHITE, ui_rect, 2)
        
        y_offset = 70
        
        # Title
        title = self.big_font.render("Traffic Control Dashboard", True, WHITE)
        self.screen.blit(title, (ui_rect.x + 10, y_offset))
        y_offset += 50
        
        # Real-time statistics
        stats = [
            f"🚗 Active Cars: {len(self.cars)}",
            f"📊 Total Spawned: {self.total_cars_spawned}",
            f"🚨 Emergency Vehicles: {self.emergency_vehicles}",
            f"⏱️ Avg Wait Time: {(sum(car.total_wait_time for car in self.cars) / max(1, len(self.cars))):.1f}s",
            f"🤖 AI Model: {'✅ Active' if self.model.is_trained else '⏳ Training...'}",
            f"🕐 Time: {self.model.current_hour:02d}:00",
            f"📅 Day: {self.model.current_day}",
        ]
        
        for stat in stats:
            text = self.font.render(stat, True, WHITE)
            self.screen.blit(text, (ui_rect.x + 10, y_offset))
            y_offset += 25
        
        y_offset += 20
        
        # Active intersections
        header = self.font.render("🚦 High Priority Intersections:", True, YELLOW)
        self.screen.blit(header, (ui_rect.x + 10, y_offset))
        y_offset += 30
        
        # Sort intersections by priority
        active_intersections = [i for i in self.intersections if i.vehicle_count > 0]
        active_intersections.sort(key=lambda x: (x.emergency_present, x.vehicle_count), reverse=True)
        
        for intersection in active_intersections[:8]:
            # Emergency indicator
            prefix = "🚨" if intersection.emergency_present else "🚗"
            color = RED if intersection.emergency_present else ORANGE if intersection.vehicle_count > 4 else YELLOW
            
            info = f"{prefix} {intersection.id}: {intersection.vehicle_count} cars"
            text = self.font.render(info, True, color)
            self.screen.blit(text, (ui_rect.x + 10, y_offset))
            
            # Light state
            light_color_text = {"red": "🔴", "yellow": "🟡", "green": "🟢"}
            light_info = f"    {light_color_text[intersection.light_state]} {intersection.action_recommendation}"
            light_text = self.font.render(light_info, True, CYAN)
            self.screen.blit(light_text, (ui_rect.x + 20, y_offset + 15))
            y_offset += 40
        
        # Controls
        y_offset = ui_rect.bottom - 200
        controls = [
            "🎮 Controls:",
            "SPACE - Pause/Resume",
            "R - Reset Simulation", 
            "T - Change Time (+1 hour)",
            "D - Change Day",
            "S - Save Data & Exit",
            "ESC - Quick Exit"
        ]
        
        for i, control in enumerate(controls):
            color = YELLOW if i == 0 else WHITE
            text = self.font.render(control, True, color)
            self.screen.blit(text, (ui_rect.x + 10, y_offset + i * 20))
    
    def handle_events(self):
        """Handle pygame events"""
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                self.save_and_exit()
            elif event.type == pygame.KEYDOWN:
                if event.key == pygame.K_SPACE:
                    self.paused = not self.paused
                elif event.key == pygame.K_r:
                    self.reset_simulation()
                elif event.key == pygame.K_t:
                    # Change time
                    self.model.current_hour = (self.model.current_hour + 1) % 24
                elif event.key == pygame.K_d:
                    # Change day
                    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                    current_index = days.index(self.model.current_day)
                    self.model.current_day = days[(current_index + 1) % len(days)]
                elif event.key == pygame.K_s:
                    self.save_and_exit()
                elif event.key == pygame.K_ESCAPE:
                    self.running = False
    
    def reset_simulation(self):
        """Reset the simulation"""
        self.cars.clear()
        self.total_cars_spawned = 0
        self.emergency_vehicles = 0
        self.frame_count = 0
        
        # Reset intersections
        for intersection in self.intersections:
            intersection.light_state = 'red'
            intersection.light_timer = 0
            intersection.cars_waiting = []
            intersection.congestion_level = 0
            intersection.vehicle_count = 0
            intersection.emergency_present = False
            intersection.total_cars_passed = 0
            intersection.avg_wait_time = 0
            intersection.light_changes = 0
        
        # Reset data logger
        self.data_logger = DataLogger()
        print("🔄 Simulation Reset")
    
    def save_and_exit(self):
        """Save data and exit"""
        print("💾 Saving simulation data...")
        
        # Log all remaining cars
        for car in self.cars:
            self.data_logger.log_car(car)
        
        # Log all intersections
        for intersection in self.intersections:
            self.data_logger.log_intersection(intersection)
        
        # Save to Excel
        success = self.data_logger.save_to_excel()
        
        if success:
            print("✅ Data saved successfully!")
        else:
            print("❌ Failed to save data")
        
        self.running = False
    
    def run(self):
        """Main simulation loop"""
        print("🚦 Starting Smart Traffic Management System...")
        print("📊 Training AI model in background...")
        
        while self.running:
            self.handle_events()
            
            if not self.paused:
                # Spawn cars periodically
                if self.frame_count % 90 == 0:  # Every 1.5 seconds at 60 FPS
                    if random.random() < 0.7:  # 70% chance to spawn
                        self.spawn_car()
                
                # Update simulation
                self.update_intersections()
                self.update_cars()
                
                # Update time simulation (1 hour per 10 seconds)
                if self.frame_count % 600 == 0:  # Every 10 seconds
                    self.model.current_hour = (self.model.current_hour + 1) % 24
                    if self.model.current_hour == 0:
                        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                        current_index = days.index(self.model.current_day)
                        self.model.current_day = days[(current_index + 1) % len(days)]
                
                self.frame_count += 1
            
            # Draw everything
            self.screen.fill(BLACK)
            self.draw_roads()
            
            # Draw intersections
            for intersection in self.intersections:
                intersection.draw(self.screen, self.font)
            
            # Draw cars
            for car in self.cars:
                car.draw(self.screen)
            
            # Draw UI
            self.draw_ui()
            
            # Show pause indicator
            if self.paused:
                pause_text = self.big_font.render("⏸️ PAUSED", True, YELLOW)
                pause_rect = pause_text.get_rect(center=(WINDOW_WIDTH // 2, 30))
                pygame.draw.rect(self.screen, BLACK, pause_rect.inflate(20, 10))
                self.screen.blit(pause_text, pause_rect)
            
            pygame.display.flip()
            self.clock.tick(60)  # 60 FPS
        
        pygame.quit()
        print("🏁 Smart Traffic Management System Closed")


# Main execution
if __name__ == "__main__":
    try:
        simulation = TrafficSimulation()
        simulation.run()
    except KeyboardInterrupt:
        print("\n⚠️ Simulation interrupted by user")
        pygame.quit()
    except Exception as e:
        print(f"❌ Error: {e}")
        pygame.quit()